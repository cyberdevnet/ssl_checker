import openpyxl
from datetime import datetime
import subprocess
import re
from urllib.parse import urlparse
import socket

# Load workbook and sheet
workbook = openpyxl.load_workbook('certificati_iso_date_with_status.xlsx')
sheet = workbook['Sheet1']

############ Find column indices #################
header = [cell.value for cell in sheet[1]]

try:
    url_col_index = header.index("URL") + 1
    cert_col = header.index("cert scadenza") + 1
    online_col = header.index("sito online") + 1
    ssl_error_col = header.index("SSL error") + 1
except ValueError as e:
    raise ValueError(f"Missing expected column: {e}")

# Add "expiration_status" column if missing
if "expiration_status" not in header:
    expiration_status_col = len(header) + 1
    sheet.cell(row=1, column=expiration_status_col).value = "expiration_status"
else:
    expiration_status_col = header.index("expiration_status") + 1

############ Extract all URLs #################
urls = []
for row in sheet.iter_rows(min_row=2, min_col=url_col_index, max_col=url_col_index, values_only=True):
    urls.append(row[0])

############ Cert verification and online check #################

def get_hostname(url):
    parsed = urlparse(url)
    return parsed.hostname

def get_cert_info(hostname, port=443):
    try:
        socket.gethostbyname(hostname)
        online_status = "yes"

        result = subprocess.run(
            ["openssl", "s_client", "-connect", f"{hostname}:{port}", "-servername", hostname, "-showcerts"],
            input="",
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=10,
            text=True
        )

        output = result.stdout
        if result.returncode != 0 or not output:
            return "Connection failed", "no", "", ""

        # Parse expiry date
        match = re.search(r"NotAfter: (.+)", output)
        if match:
            expiry_str = match.group(1).strip()
            try:
                expiry_dt = datetime.strptime(expiry_str, "%b %d %H:%M:%S %Y GMT").date()
                today = datetime.today().date()
                days_left = (expiry_dt - today).days
                expiry_date_str = expiry_dt.isoformat()

                if days_left < 0:
                    expiry_warn = f"❌ EXPIRED {abs(days_left)} days ago"
                elif days_left <= 30:
                    expiry_warn = f"⚠️ Expires in {days_left} days"
                else:
                    expiry_warn = ""
            except ValueError:
                expiry_date_str = f"Unparsable date: {expiry_str}"
                expiry_warn = "❌ Invalid expiry format"
        else:
            expiry_date_str = "NotAfter not found"
            expiry_warn = ""

        # Extract all verify errors
        verify_errors = re.findall(r"verify error:[^\n]+", result.stderr)
        ssl_error_msg = "; ".join(verify_errors) if verify_errors else ""

        return expiry_date_str, online_status, ssl_error_msg, expiry_warn

    except socket.gaierror:
        return "DNS resolution failed", "no", "", ""
    except subprocess.TimeoutExpired:
        return "Timeout", "no", "", ""
    except Exception as e:
        return f"Error: {str(e)}", "no", "", ""

############ Process each URL #################

expiration_issues = []
ssl_errors_list = []

for i, url in enumerate(urls, start=2):
    hostname = get_hostname(url)
    if hostname:
        expiry, online, ssl_error, expiration_status = get_cert_info(hostname)
    else:
        expiry, online, ssl_error, expiration_status = "Invalid URL", "no", "", ""

    sheet.cell(row=i, column=cert_col).value = expiry
    sheet.cell(row=i, column=online_col).value = online
    sheet.cell(row=i, column=ssl_error_col).value = ssl_error
    sheet.cell(row=i, column=expiration_status_col).value = expiration_status

    print(f"{hostname} => {expiry}, online: {online}, ssl error: {ssl_error or 'none'}, expiration status: {expiration_status or 'none'}")

    # Collect summary info
    if expiration_status:
        expiration_issues.append((hostname, expiration_status))
    if ssl_error:
        ssl_errors_list.append((hostname, ssl_error))

# Save workbook
workbook.save("certificati_iso_date_with_status.xlsx")

############ Print summaries #############

print("\nSummary of sites with expiration warnings/errors:")
if expiration_issues:
    for host, status in expiration_issues:
        print(f" - {host}: {status}")
else:
    print(" None")

print("\nSummary of sites with SSL errors:")
if ssl_errors_list:
    for host, error in ssl_errors_list:
        print(f" - {host}: {error}")
else:
    print(" None")


summary_lines = []

summary_lines.append("Summary of sites with expiration warnings/errors:")
if expiration_issues:
    for host, status in expiration_issues:
        summary_lines.append(f" - {host}: {status}")
else:
    summary_lines.append(" None")

summary_lines.append("\nSummary of sites with SSL errors:")
if ssl_errors_list:
    for host, error in ssl_errors_list:
        summary_lines.append(f" - {host}: {error}")
else:
    summary_lines.append(" None")

# Save summary to a file
with open("summary_report.txt", "w") as f:
    f.write("\n".join(summary_lines))
